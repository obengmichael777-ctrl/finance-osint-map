# tests/generate_test_data_v2.py
"""
Pan-Asian Retail Test Data Generator
Generates test data compatible with the schema-driven extraction pipeline.

Usage:
    python tests/generate_test_data_v2.py
    python tests/generate_test_data_v2.py --stores 5 --rows 200
"""

import pandas as pd
import numpy as np
from faker import Faker
from datetime import datetime, timedelta
import random
import os
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import argparse
import yaml

# Initialize Faker instances for each locale
FAKER_LOCALES = {
    'zh_CN': Faker('zh_CN'),  # Chinese
    'ja_JP': Faker('ja_JP'),  # Japanese
    'ko_KR': Faker('ko_KR'),  # Korean
    'th_TH': Faker('th_TH'),  # Thai
    'vi_VN': Faker('vi_VN'),  # Vietnamese
    'en_MS': Faker('en_MS'),  # Malay
    'id_ID': Faker('id_ID')  # Indonesian
    #'en_HK': Faker('en_HK'),  # Hong Kong English - Deprecated in new faker versions
}

# Default English faker for fallback
fake_en = Faker()


class PanAsianStoreDataGenerator:
    """
    Generate realistic Pan-Asian supermarket data compatible with
    the schema-driven extraction pipeline.

    Key design decisions for pipeline compatibility:
    1. Uses expected column names from schema_registry
    2. Generates proper data types (not all strings)
    3. Creates store metadata with lat/lon coordinates
    4. Produces files named according to schema patterns
    """

    def __init__(self, output_dir="tests/fixtures/test_data"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Region definitions with real coordinates
        self.regions = {
            'JP': {
                'country': 'JP',
                'country_name': 'Japan',
                'region': 'East Asia Developed',
                'currency': 'JPY',
                'locale': 'ja_JP',
                'cities': [
                    ('Tokyo', 35.6762, 139.6503),
                    ('Osaka', 34.6937, 135.5023),
                    ('Nagoya', 35.1815, 136.9066),
                    ('Fukuoka', 33.5904, 130.4017),
                    ('Sapporo', 43.0618, 141.3545),
                    ('Yokohama', 35.4437, 139.6380),
                ]
            },
            'CN': {
                'country': 'CN',
                'country_name': 'China',
                'region': 'Greater China',
                'currency': 'CNY',
                'locale': 'zh_CN',
                'cities': [
                    ('Shanghai', 31.2304, 121.4737),
                    ('Beijing', 39.9042, 116.4074),
                    ('Guangzhou', 23.1291, 113.2644),
                    ('Shenzhen', 22.5431, 114.0579),
                    ('Chengdu', 30.5728, 104.0668),
                    ('Hangzhou', 30.2741, 120.1551),
                ]
            },
            'KR': {
                'country': 'KR',
                'country_name': 'South Korea',
                'region': 'East Asia Developed',
                'currency': 'KRW',
                'locale': 'ko_KR',
                'cities': [
                    ('Seoul', 37.5665, 126.9780),
                    ('Busan', 35.1796, 129.0756),
                    ('Incheon', 37.4563, 126.7052),
                    ('Daegu', 35.8722, 128.6025),
                ]
            },
            'HK': {
                'country': 'HK',
                'country_name': 'Hong Kong',
                'region': 'Greater China',
                'currency': 'HKD',
                'locale': 'en_HK',
                'cities': [
                    ('Central', 22.2797, 114.1591),
                    ('Kowloon', 22.3193, 114.1694),
                    ('Tsim Sha Tsui', 22.2973, 114.1719),
                    ('Causeway Bay', 22.2807, 114.1849),
                ]
            },
            'TH': {
                'country': 'TH',
                'country_name': 'Thailand',
                'region': 'ASEAN Core',
                'currency': 'THB',
                'locale': 'th_TH',
                'cities': [
                    ('Bangkok', 13.7563, 100.5018),
                    ('Chiang Mai', 18.7883, 98.9853),
                    ('Phuket', 7.8804, 98.3923),
                ]
            },
            'VN': {
                'country': 'VN',
                'country_name': 'Vietnam',
                'region': 'ASEAN Frontier',
                'currency': 'VND',
                'locale': 'vi_VN',
                'cities': [
                    ('Hanoi', 21.0278, 105.8342),
                    ('Ho Chi Minh City', 10.8231, 106.6297),
                    ('Da Nang', 16.0544, 108.2022),
                ]
            },
            'MY': {
                'country': 'MY',
                'country_name': 'Malaysia',
                'region': 'ASEAN Core',
                'currency': 'MYR',
                'locale': 'ms_MY',
                'cities': [
                    ('Kuala Lumpur', 3.1390, 101.6869),
                    ('Penang', 5.4141, 100.3288),
                    ('Johor Bahru', 1.4927, 103.7414),
                ]
            },
            'ID': {
                'country': 'ID',
                'country_name': 'Indonesia',
                'region': 'ASEAN Core',
                'currency': 'IDR',
                'locale': 'id_ID',
                'cities': [
                    ('Jakarta', -6.2088, 106.8456),
                    ('Surabaya', -7.2575, 112.7521),
                    ('Bandung', -6.9175, 107.6191),
                    ('Bali', -8.3405, 115.0920),
                ]
            },
        }

        # Product categories with local language names
        self.product_categories = {
            'zh_CN': ['饮料', '零食', '水果', '蔬菜', '肉类', '海鲜', '乳制品', '调味品'],
            'ja_JP': ['飲料', 'スナック', '果物', '野菜', '肉', '海鮮', '乳製品', '調味料'],
            'ko_KR': ['음료', '스낵', '과일', '채소', '고기', '해산물', '유제품', '조미료'],
            'th_TH': ['เครื่องดื่ม', 'ขนม', 'ผลไม้', 'ผัก', 'เนื้อ', 'ทะเล', 'นม', 'เครื่องปรุง'],
            'vi_VN': ['đồ uống', 'đồ ăn vặt', 'trái cây', 'rau', 'thịt', 'hải sản', 'sữa', 'gia vị'],
            'ms_MY': ['minuman', 'snek', 'buah', 'sayur', 'daging', 'makanan laut', 'tenusu', 'perasa'],
            'id_ID': ['minuman', 'camilan', 'buah', 'sayur', 'daging', 'seafood', 'susu', 'bumbu'],
            'en_HK': ['Beverages', 'Snacks', 'Fruit', 'Vegetables', 'Meat', 'Seafood', 'Dairy', 'Condiments'],
        }

        self.store_metadata = []

    def _get_faker(self, locale_code):
        """Get appropriate Faker instance for locale"""
        return FAKER_LOCALES.get(locale_code, fake_en)

    def _generate_store_id(self, country_code, store_num):
        """Generate standardized store ID"""
        return f"store_{country_code}_{store_num:03d}"

    def _generate_product_id(self):
        """Generate standardized product ID (SKU)"""
        return f"SKU-{random.randint(10000, 99999)}"

    def _jitter_coordinates(self, lat, lon, radius_km=5):
        """
        Add random jitter to coordinates within radius.
        Simulates stores within same city area.
        """
        # 1 degree ≈ 111 km
        lat_jitter = random.uniform(-radius_km/111, radius_km/111)
        lon_jitter = random.uniform(-radius_km/111, radius_km/111)
        return lat + lat_jitter, lon + lon_jitter

    def create_store_data(self, region_code, store_num, num_rows=100):
        """
        Generate data for a single store.

        Creates data matching schema_registry expectations:
        - Sheet named 'Sales_2024' or 'Daily_Sales_2024'
        - Columns: 'Transaction Date', 'SKU', 'Qty Sold', 'Net Sales', etc.
        """
        region_info = self.regions[region_code]
        locale = region_info['locale']
        currency = region_info['currency']

        # Select a city for this store
        city_name, city_lat, city_lon = random.choice(region_info['cities'])
        store_lat, store_lon = self._jitter_coordinates(city_lat, city_lon)

        # Generate store ID
        store_id = self._generate_store_id(region_code, store_num)

        # Record store metadata for schema registry
        self.store_metadata.append({
            'store_id': store_id,
            'country': region_code,
            'country_name': region_info['country_name'],
            'region': region_info['region'],
            'currency': currency,
            'lat': round(store_lat, 6),
            'lon': round(store_lon, 6),
            'city': city_name,
            'file_pattern': f"{store_id}_sales_*.xlsx",
        })

        # Generate transaction dates (last 12 months, some gaps for realism)
        end_date = datetime.now()
        start_date = end_date - timedelta(days=365)

        # Generate dates with intentional gaps (weekends missing for some)
        all_dates = pd.date_range(start=start_date, end=end_date, freq='D')

        # Simulate: store might be closed some days
        operating_days = [d for d in all_dates if random.random() > 0.05]  # 95% open rate

        # Select num_rows random operating days
        if len(operating_days) > num_rows:
            selected_dates = sorted(random.sample(list(operating_days), num_rows))
        else:
            selected_dates = list(operating_days)
            num_rows = len(selected_dates)

        # Generate transaction data
        data = []
        for date in selected_dates:
            # Multiple transactions per day
            num_transactions = random.randint(1, 10)

            for _ in range(num_transactions):
                product_id = self._generate_product_id()
                qty = random.randint(1, 30)
                unit_price = round(random.uniform(1.0, 500.0), 2)
                net_sales = round(qty * unit_price * random.uniform(0.9, 1.1), 2)

                # Some realistic negative scenarios (returns)
                if random.random() < 0.02:  # 2% chance of return
                    qty = -qty
                    net_sales = -net_sales

                data.append({
                    'Transaction Date': date,  # DateTime object, not string
                    'SKU': product_id,
                    'Product Name': self._generate_localized_product(locale),
                    'Category': random.choice(self.product_categories.get(locale, ['General'])),
                    'Qty Sold': qty,
                    'Unit Price': unit_price,
                    'Net Sales': round(net_sales, 2),
                    'Customer ID': f'CUST-{random.randint(1000, 9999)}' if random.random() > 0.1 else None,
                    'Payment Type': random.choice(['CASH', 'CREDIT', 'DEBIT', 'MOBILE', 'E-WALLET']),
                })

        df = pd.DataFrame(data)

        # Add store metadata columns
        df['Store ID'] = store_id
        df['Country'] = region_code
        df['City'] = city_name

        # Generate store name in local language
        faker = self._get_faker(locale)
        df['Store Name'] = f"{faker.company()} {random.choice(['Supermarket', 'Store', 'Mart'])}"

        # Ensure proper data types
        df['Transaction Date'] = pd.to_datetime(df['Transaction Date'])
        df['Qty Sold'] = df['Qty Sold'].astype(int)
        df['Net Sales'] = df['Net Sales'].astype(float)
        df['Unit Price'] = df['Unit Price'].astype(float)
        df['SKU'] = df['SKU'].astype(str)
        df['Customer ID'] = df['Customer ID'].astype(str).replace('nan', 'UNKNOWN')

        return df, store_id, region_info

    def _generate_localized_product(self, locale):
        """Generate product name in local language"""
        faker = self._get_faker(locale)

        templates = {
            'zh_CN': lambda: f"{faker.word()}牌{random.choice(['牛奶', '面包', '饼干', '饮料', '零食'])}",
            'ja_JP': lambda: f"{faker.word()}の{random.choice(['お茶', 'お菓子', 'パン', '飲料', '麺類'])}",
            'ko_KR': lambda: f"{faker.word()} {random.choice(['우유', '빵', '과자', '음료', '라면'])}",
            'en_HK': lambda: f"{faker.word()} {random.choice(['Milk', 'Bread', 'Biscuit', 'Drink', 'Noodle'])}",
        }

        template = templates.get(locale, lambda: f"Product {faker.ean8()}")
        return template()

    def save_store_excel(self, df, store_id, region_info, format_type='standard'):
        """
        Save store data as Excel file matching schema expectations.

        Supports multiple formats to test schema flexibility:
        - 'standard': Sheet named 'Sales_2024', standard columns
        - 'uk_format': Sheet named 'Daily_Sales_2024', UK date format
        - 'localized': Sheet name in local language, varied columns
        """
        store_dir = self.output_dir / store_id
        store_dir.mkdir(parents=True, exist_ok=True)

        # Determine file naming based on schema patterns
        timestamp = datetime.now().strftime('%Y%m%d')
        filename = f"{store_id}_sales_{timestamp}.xlsx"
        filepath = store_dir / filename

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            workbook = writer.book

            # Sales sheet with proper naming
            if format_type == 'uk_format':
                sheet_name = f"Daily_Sales_{datetime.now().year}"
            elif format_type == 'localized':
                locale_names = {
                    'zh_CN': f"销售_{datetime.now().year}",
                    'ja_JP': f"売上_{datetime.now().year}",
                }
                sheet_name = locale_names.get(region_info['locale'], f"Sales_{datetime.now().year}")
            else:
                sheet_name = f"Sales_{datetime.now().year}"

            # Add header rows (some stores have title rows, testing skip_rows)
            sales_sheet = workbook.create_sheet(sheet_name, 0)

            # Title row (simulates real-world formatting)
            if random.random() > 0.5:
                sales_sheet.cell(row=1, column=1, value=f"Store Sales Report - {store_id}")
                sales_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
                sales_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

                sales_sheet.cell(row=2, column=1, value=f"Period: Last 12 Months")
                sales_sheet.cell(row=2, column=1).font = Font(italic=True, size=10)

                start_row = 2  # Data starts after title rows
            else:
                start_row = 0

            # Write data
            df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)

            # Adjust column widths
            worksheet = writer.sheets[sheet_name]
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).str.len().max(),
                    len(str(col))
                )
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 30)

            # Category summary sheet (bonus sheet for testing multi-sheet handling)
            if 'Category' in df.columns:
                cat_summary = df.groupby('Category').agg({
                    'Net Sales': ['sum', 'count'],
                    'Qty Sold': 'sum'
                }).round(2)
                cat_summary.to_excel(writer, sheet_name='Category_Summary')

        return filepath

    def generate_all_test_data(self, stores_per_region=3, transactions_per_store=150):
        """
        Generate complete test dataset across all regions.

        Args:
            stores_per_region: Number of stores per country
            transactions_per_store: Approximate transactions per store
        """
        generated_files = []
        all_data = []

        print("=" * 60)
        print("GENERATING PAN-ASIAN RETAIL TEST DATA")
        print("=" * 60)

        for region_code, region_info in self.regions.items():
            print(f"\n📍 {region_info['country_name']} ({region_code})")

            for store_num in range(1, stores_per_region + 1):
                # Generate data
                df, store_id, info = self.create_store_data(
                    region_code, store_num, transactions_per_store
                )

                # Save in different formats to test schema flexibility
                format_type = random.choice(['standard', 'standard', 'uk_format', 'standard'])
                filepath = self.save_store_excel(df, store_id, info, format_type)

                generated_files.append(filepath)
                all_data.append(df)

                print(f"  ✅ {store_id}: {len(df)} transactions -> {filepath.name}")

        # Save store metadata for schema registry
        metadata_path = self.output_dir / 'store_metadata.json'
        with open(metadata_path, 'w') as f:
            json.dump(self.store_metadata, f, indent=2, default=str)
        print(f"\n📄 Store metadata saved to {metadata_path}")

        # Save combined dataset for direct testing
        if all_data:
            combined = pd.concat(all_data, ignore_index=True)
            combined_path = self.output_dir / 'combined_test_data.parquet'
            combined.to_parquet(combined_path, compression='snappy')
            print(f"📦 Combined dataset: {len(combined)} rows -> {combined_path}")

        # Generate YAML schema config based on generated data
        self.generate_schema_config()

        print(f"\n{'=' * 60}")
        print(f"✅ Generated {len(generated_files)} Excel files across {len(self.regions)} countries")
        print(f"✅ Total stores: {len(self.store_metadata)}")
        print(f"✅ Total transactions: {sum(len(d) for d in all_data)}")
        print(f"{'=' * 60}")

        return generated_files, self.store_metadata

    def generate_schema_config(self):
        """
        Automatically generate schema configuration from test data.
        This ensures the schema registry can find and process generated files.
        """
        schema_config = {'stores': {}}

        for store in self.store_metadata:
            store_id = store['store_id']

            schema_config['stores'][store_id] = {
                'country': store['country'],
                'region': store['region'],
                'lat': store['lat'],
                'lon': store['lon'],
                'file_pattern': f"{store_id}_sales_*.xlsx",
                'file_metadata': {
                    'timezone': 'Asia/Tokyo' if store['country'] == 'JP' else 'Asia/Shanghai',
                    'currency': store['currency'],
                },
                'sheets': {
                    'sales_transactions': {
                        'sheet_name': f"regex:^(Sales_|Daily_Sales_|销售_|売上_)",
                        'skip_rows': 0,  # Will be auto-detected
                        'header_row': 0,
                        'columns': {
                            'date': {
                                'source_name': 'Transaction Date',
                                'data_type': 'date',
                                'required': True,
                                'validation_rules': {'not_null': True}
                            },
                            'product_id': {
                                'source_name': 'SKU',
                                'data_type': 'string',
                                'required': True
                            },
                            'qty': {
                                'source_name': 'Qty Sold',
                                'data_type': 'integer',
                                'required': True,
                                'validation_rules': {'min': -100, 'max': 100000}
                            },
                            'revenue': {
                                'source_name': 'Net Sales',
                                'data_type': 'float',
                                'required': True,
                                'validation_rules': {'min': -50000}
                            },
                            'customer_id': {
                                'source_name': 'Customer ID',
                                'data_type': 'string',
                                'required': False,
                                'default_value': 'UNKNOWN'
                            },
                            'payment_method': {
                                'source_name': 'Payment Type',
                                'data_type': 'category',
                                'required': False,
                                'validation_rules': {
                                    'allowed_values': [
                                        'CASH', 'CREDIT', 'DEBIT', 'MOBILE', 'E-WALLET'
                                    ]
                                }
                            }
                        }
                    }
                }
            }

        # Save schema config
        config_path = Path('config/store_schemas.yaml')
        config_path.parent.mkdir(parents=True, exist_ok=True)

        with open(config_path, 'w') as f:
            yaml.dump(schema_config, f, default_flow_style=False, allow_unicode=True)

        print(f"📋 Schema config generated: {config_path}")
        return config_path

    def create_edge_case_files(self):
        """Create intentionally problematic files for testing DLQ"""
        edge_cases_dir = self.output_dir / 'edge_cases'
        edge_cases_dir.mkdir(exist_ok=True)

        print("\n🔧 Generating edge case files...")

        # 1. Empty file
        empty_path = edge_cases_dir / 'empty_file.xlsx'
        wb = Workbook()
        wb.save(empty_path)
        print(f"  ✅ Empty file: {empty_path}")

        # 2. Missing required columns
        bad_df = pd.DataFrame({
            'Wrong Date': ['2024-01-01'],
            'Product': ['TEST'],
            'Amount': [100.00]
        })
        bad_path = edge_cases_dir / 'missing_columns.xlsx'
        bad_df.to_excel(bad_path, sheet_name='Sales_2024', index=False)
        print(f"  ✅ Missing columns: {bad_path}")

        # 3. File with negative revenue (tests range validation)
        neg_df = pd.DataFrame({
            'Transaction Date': pd.date_range('2024-01-01', periods=20),
            'SKU': ['SKU-TEST'] * 20,
            'Qty Sold': [5, -3, 10, -2] * 5,
            'Net Sales': [100, -50, 200, -75] * 5,
            'Customer ID': ['CUST-001'] * 20,
            'Payment Type': ['CASH'] * 20
        })
        neg_path = edge_cases_dir / 'negative_values.xlsx'
        neg_df.to_excel(neg_path, sheet_name='Sales_2024', index=False)
        print(f"  ✅ Negative values: {neg_path}")

        # 4. Corrupted file
        corrupt_path = edge_cases_dir / 'corrupted.xlsx'
        with open(corrupt_path, 'wb') as f:
            f.write(b'This is not an Excel file' + b'\x00' * 100)
        print(f"  ✅ Corrupted file: {corrupt_path}")

        return edge_cases_dir


def main():
    parser = argparse.ArgumentParser(description='Generate Pan-Asian retail test data')
    parser.add_argument('--stores', type=int, default=2, help='Stores per region')
    parser.add_argument('--rows', type=int, default=150, help='Transactions per store')
    parser.add_argument('--output', type=str, default='tests/fixtures', help='Output directory')
    parser.add_argument('--edge-cases', action='store_true', help='Generate edge case files')

    args = parser.parse_args()

    generator = PanAsianStoreDataGenerator(output_dir=args.output)

    # Generate main test data
    files, metadata = generator.generate_all_test_data(
        stores_per_region=args.stores,
        transactions_per_store=args.rows
    )

    # Generate edge cases if requested
    if args.edge_cases:
        generator.create_edge_case_files()

    print(f"\n🎯 Next steps:")
    print(f"1. Run: python run_pipeline.py --mode batch --initial-dir {args.output}")
    print(f"2. Start API: uvicorn api:app --reload --port 8000")
    print(f"3. Visit: http://localhost:8000/api/v1/markers")


if __name__ == '__main__':
    main()
